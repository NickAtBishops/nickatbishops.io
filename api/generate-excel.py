from http.server import BaseHTTPRequestHandler
import json
import io
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import base64
import os

# Path to template (will be in the same deployment)
TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), '..', 'templates', 'GGC_UW_Template.xlsx')

class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        try:
            # Read request body
            content_length = int(self.headers['Content-Length'])
            body = self.rfile.read(content_length)
            data = json.loads(body)
            
            # Generate Excel
            excel_bytes = generate_excel(data)
            
            # Send response
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', 'attachment; filename="GGC_Underwriting.xlsx"')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(excel_bytes)
            
        except Exception as e:
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps({'error': str(e)}).encode())
    
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()


def generate_excel(data):
    """Generate formatted Excel from JSON data using template"""
    
    # Load template or create new workbook
    if os.path.exists(TEMPLATE_PATH):
        wb = load_workbook(TEMPLATE_PATH)
    else:
        wb = create_workbook_from_scratch(data)
        
    # Fill data into template
    fill_deal_summary(wb, data)
    fill_income_expenses(wb, data)
    fill_rent_roll(wb, data)
    fill_rent_comps(wb, data)
    fill_sale_comps(wb, data)
    fill_market_analysis(wb, data)
    fill_flags_questions(wb, data)
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def create_workbook_from_scratch(data):
    """Create a new formatted workbook if template doesn't exist"""
    from openpyxl import Workbook
    
    wb = Workbook()
    
    # Define styles
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    subheader_fill = PatternFill(start_color="D6E3F8", end_color="D6E3F8", fill_type="solid")
    subheader_font = Font(bold=True, size=11)
    currency_format = '$#,##0'
    percent_format = '0.0%'
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Store styles for use in fill functions
    wb._styles = {
        'header_fill': header_fill,
        'header_font': header_font,
        'subheader_fill': subheader_fill,
        'subheader_font': subheader_font,
        'currency_format': currency_format,
        'percent_format': percent_format,
        'thin_border': thin_border
    }
    
    # Create sheets
    sheets = ['Deal Summary', 'Income & Expenses', 'Rent Roll', 'Rent Comps', 
              'Sale Comps', 'Market Analysis', 'Flags & Questions', 'Pro Forma', 'Sources & Uses']
    
    # Rename first sheet
    wb.active.title = sheets[0]
    
    # Create remaining sheets
    for sheet_name in sheets[1:]:
        wb.create_sheet(sheet_name)
    
    return wb


def fill_deal_summary(wb, data):
    """Fill Deal Summary sheet"""
    if 'Deal Summary' not in wb.sheetnames:
        return
    
    ws = wb['Deal Summary']
    fin = data.get('financials', {})
    prop = fin.get('propertyInfo', {}) or data.get('propertyInfo', {})
    
    # Property info
    ws['B4'] = prop.get('name', '')
    ws['B5'] = prop.get('address', '')
    ws['B6'] = f"{prop.get('city', '')}, {prop.get('state', '')}"
    ws['B7'] = prop.get('totalUnits', 0)
    ws['B8'] = prop.get('askingPrice', 0) or data.get('propertyInfo', {}).get('askingPrice', 0)
    
    # Calculate metrics
    income_total = sum(i.get('t12Amount', 0) or 0 for i in fin.get('income', []))
    expense_total = sum(e.get('t12Amount', 0) or 0 for e in fin.get('expenses', []))
    noi = income_total - expense_total
    
    ws['B12'] = noi
    ws['B13'] = noi  # GGC normalized (simplified)
    
    asking = prop.get('askingPrice', 0) or data.get('propertyInfo', {}).get('askingPrice', 0)
    ws['B14'] = noi / asking if asking else 0
    ws['B15'] = expense_total / income_total if income_total else 0
    
    rr = fin.get('rentRoll', {})
    ws['B16'] = rr.get('occupancyRate', 0)


def fill_income_expenses(wb, data):
    """Fill Income & Expenses sheet"""
    if 'Income & Expenses' not in wb.sheetnames:
        return
    
    ws = wb['Income & Expenses']
    fin = data.get('financials', {})
    
    # Headers
    headers = ['Seller Line Item', 'GGC Category', 'T12 Amount', 'T3 Amount', 'Confidence', 'Notes']
    
    # Style setup
    styles = getattr(wb, '_styles', {})
    header_fill = styles.get('header_fill', PatternFill(start_color="1F4E79", fill_type="solid"))
    header_font = styles.get('header_font', Font(color="FFFFFF", bold=True))
    
    # Income section
    ws['A1'] = 'INCOME'
    ws['A1'].font = Font(bold=True, size=14)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    row = 4
    for item in fin.get('income', []):
        ws.cell(row=row, column=1, value=item.get('sellerName', ''))
        ws.cell(row=row, column=2, value=item.get('ggcCategory', ''))
        ws.cell(row=row, column=3, value=item.get('t12Amount', 0))
        ws.cell(row=row, column=4, value=item.get('t3Amount', 0))
        ws.cell(row=row, column=5, value=item.get('confidence', ''))
        ws.cell(row=row, column=6, value=item.get('notes', ''))
        row += 1
    
    # Expense section
    row += 2
    ws.cell(row=row, column=1, value='EXPENSES')
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    row += 2
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    row += 1
    
    for item in fin.get('expenses', []):
        ws.cell(row=row, column=1, value=item.get('sellerName', ''))
        ws.cell(row=row, column=2, value=item.get('ggcCategory', ''))
        ws.cell(row=row, column=3, value=item.get('t12Amount', 0))
        ws.cell(row=row, column=4, value=item.get('t3Amount', 0))
        ws.cell(row=row, column=5, value=item.get('confidence', ''))
        ws.cell(row=row, column=6, value=item.get('notes', ''))
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 40


def fill_rent_roll(wb, data):
    """Fill Rent Roll sheet"""
    if 'Rent Roll' not in wb.sheetnames:
        return
    
    ws = wb['Rent Roll']
    rr = data.get('financials', {}).get('rentRoll', {})
    
    ws['A1'] = 'RENT ROLL SUMMARY'
    ws['A1'].font = Font(bold=True, size=14)
    
    ws['A3'] = 'Total Units'
    ws['B3'] = rr.get('totalUnits', 0)
    ws['A4'] = 'Occupied Units'
    ws['B4'] = rr.get('occupiedUnits', 0)
    ws['A5'] = 'Vacant Units'
    ws['B5'] = rr.get('vacantUnits', 0)
    ws['A6'] = 'Occupancy Rate'
    ws['B6'] = rr.get('occupancyRate', 0)
    ws['A7'] = 'Avg Monthly Rent'
    ws['B7'] = rr.get('avgMonthlyRent', 0)
    
    # Unit mix
    ws['A9'] = 'UNIT MIX'
    ws['A9'].font = Font(bold=True, size=12)
    
    headers = ['Unit Type', 'Count', 'Occupied', 'Avg Rent']
    for col, header in enumerate(headers, 1):
        ws.cell(row=10, column=col, value=header)
        ws.cell(row=10, column=col).font = Font(bold=True)
    
    row = 11
    for unit in rr.get('unitMix', []):
        ws.cell(row=row, column=1, value=unit.get('unitType', ''))
        ws.cell(row=row, column=2, value=unit.get('count', 0))
        ws.cell(row=row, column=3, value=unit.get('occupied', 0))
        ws.cell(row=row, column=4, value=unit.get('avgRent', 0))
        row += 1


def fill_rent_comps(wb, data):
    """Fill Rent Comps sheet"""
    if 'Rent Comps' not in wb.sheetnames:
        return
    
    ws = wb['Rent Comps']
    comps = data.get('comps', {})
    
    ws['A1'] = 'RENT COMPARABLES'
    ws['A1'].font = Font(bold=True, size=14)
    
    headers = ['Property', 'Location', 'Units', 'Lot Rent', 'Occupancy', 'Distance', 'Source']
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
        ws.cell(row=3, column=col).font = Font(bold=True)
    
    row = 4
    for comp in comps.get('rentComps', []):
        ws.cell(row=row, column=1, value=comp.get('name', ''))
        ws.cell(row=row, column=2, value=f"{comp.get('city', '')}, {comp.get('state', '')}")
        ws.cell(row=row, column=3, value=comp.get('units', 0))
        ws.cell(row=row, column=4, value=comp.get('lotRent', 0))
        ws.cell(row=row, column=5, value=comp.get('occupancy', 0))
        ws.cell(row=row, column=6, value=comp.get('distance', ''))
        ws.cell(row=row, column=7, value=comp.get('source', ''))
        row += 1
    
    row += 1
    ws.cell(row=row, column=1, value='Market Rent Conclusion:')
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2, value=comps.get('marketRentConclusion', 0))


def fill_sale_comps(wb, data):
    """Fill Sale Comps sheet"""
    if 'Sale Comps' not in wb.sheetnames:
        return
    
    ws = wb['Sale Comps']
    comps = data.get('comps', {})
    
    ws['A1'] = 'SALE COMPARABLES'
    ws['A1'].font = Font(bold=True, size=14)
    
    headers = ['Property', 'Location', 'Sale Date', 'Units', 'Sale Price', '$/Unit', 'Cap Rate', 'Source']
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
        ws.cell(row=3, column=col).font = Font(bold=True)
    
    row = 4
    for comp in comps.get('saleComps', []):
        ws.cell(row=row, column=1, value=comp.get('name', ''))
        ws.cell(row=row, column=2, value=comp.get('location', ''))
        ws.cell(row=row, column=3, value=comp.get('saleDate', ''))
        ws.cell(row=row, column=4, value=comp.get('units', 0))
        ws.cell(row=row, column=5, value=comp.get('salePrice', 0))
        ws.cell(row=row, column=6, value=comp.get('pricePerUnit', 0))
        ws.cell(row=row, column=7, value=comp.get('capRate', 0))
        ws.cell(row=row, column=8, value=comp.get('source', ''))
        row += 1
    
    row += 1
    ws.cell(row=row, column=1, value='Market Cap Rate:')
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2, value=comps.get('marketCapRateConclusion', 0))


def fill_market_analysis(wb, data):
    """Fill Market Analysis sheet"""
    if 'Market Analysis' not in wb.sheetnames:
        return
    
    ws = wb['Market Analysis']
    market = data.get('market', {})
    demo = market.get('demographics', {})
    alt = market.get('altHousing', {})
    
    ws['A1'] = 'MARKET & ALTERNATIVE HOUSING ANALYSIS'
    ws['A1'].font = Font(bold=True, size=14)
    
    ws['A3'] = 'DEMOGRAPHICS'
    ws['A3'].font = Font(bold=True, size=12)
    
    ws['A4'] = 'County Population'
    ws['B4'] = demo.get('countyPopulation', 0)
    ws['A5'] = 'Population Growth (5yr)'
    ws['B5'] = demo.get('populationGrowth', 0)
    ws['A6'] = 'Median HH Income'
    ws['B6'] = demo.get('medianHHIncome', 0)
    ws['A7'] = 'Unemployment Rate'
    ws['B7'] = demo.get('unemploymentRate', 0)
    ws['A8'] = 'Major Employers'
    ws['B8'] = ', '.join(demo.get('majorEmployers', []))
    
    ws['A10'] = 'ALTERNATIVE HOUSING'
    ws['A10'].font = Font(bold=True, size=12)
    
    ws['A11'] = 'Avg SF Home Price'
    ws['B11'] = alt.get('avgSFHomePrice', 0)
    ws['A12'] = 'Avg 1BR Rent'
    ws['B12'] = alt.get('avgRent1BR', 0)
    ws['A13'] = 'Avg 2BR Rent'
    ws['B13'] = alt.get('avgRent2BR', 0)
    ws['A14'] = 'MHP All-In Cost'
    ws['B14'] = alt.get('mhpAllInCost', 0)
    
    ws['A16'] = 'Demand Signal'
    ws['B16'] = market.get('demandSignal', '')
    ws['A17'] = 'Rationale'
    ws['B17'] = market.get('demandRationale', '')


def fill_flags_questions(wb, data):
    """Fill Flags & Questions sheet"""
    if 'Flags & Questions' not in wb.sheetnames:
        return
    
    ws = wb['Flags & Questions']
    fin = data.get('financials', {})
    
    ws['A1'] = 'DILIGENCE FLAGS'
    ws['A1'].font = Font(bold=True, size=14)
    
    headers = ['#', 'Item', 'Issue', 'Severity', 'Recommendation']
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
        ws.cell(row=3, column=col).font = Font(bold=True)
    
    row = 4
    for i, flag in enumerate(fin.get('flags', []), 1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=flag.get('item', ''))
        ws.cell(row=row, column=3, value=flag.get('issue', ''))
        ws.cell(row=row, column=4, value=flag.get('severity', ''))
        ws.cell(row=row, column=5, value=flag.get('recommendation', ''))
        row += 1
    
    row += 2
    ws.cell(row=row, column=1, value='QUESTIONS FOR SELLER/BROKER')
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    row += 2
    
    for i, q in enumerate(fin.get('questions', []), 1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=q)
        row += 1
    
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 40
